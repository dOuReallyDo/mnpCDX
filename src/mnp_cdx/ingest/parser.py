"""MNP XLSX parser for monthly and daily sheets."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import hashlib
import re
from typing import Optional

import openpyxl


@dataclass
class ParseSummary:
    filename: str
    checksum: str
    monthly_records: int
    daily_records: int
    warnings: list[str]


class MNPParser:
    MONTH_ALIASES = {
        "JAN": 1,
        "FEB": 2,
        "MAR": 3,
        "APR": 4,
        "MAY": 5,
        "JUN": 6,
        "JUL": 7,
        "AUG": 8,
        "SEP": 9,
        "OCT": 10,
        "NOV": 11,
        "DEC": 12,
        "GEN": 1,
        "MAG": 5,
        "GIU": 6,
        "LUG": 7,
        "AGO": 8,
        "SET": 9,
        "OTT": 10,
        "DIC": 12,
    }

    def __init__(self, include_self_flows: bool = False) -> None:
        self.include_self_flows = include_self_flows

    @staticmethod
    def checksum(file_path: str | Path) -> str:
        sha = hashlib.sha256()
        with open(file_path, "rb") as fh:
            for chunk in iter(lambda: fh.read(4096), b""):
                sha.update(chunk)
        return sha.hexdigest()

    def parse(self, file_path: str | Path, file_id: str) -> tuple[list[dict], ParseSummary]:
        file_path = str(file_path)
        monthly, monthly_warnings = self.parse_monthly(file_path, file_id)
        daily, daily_warnings = self.parse_daily(file_path, file_id)
        all_warnings = monthly_warnings + daily_warnings
        summary = ParseSummary(
            filename=Path(file_path).name,
            checksum=self.checksum(file_path),
            monthly_records=len(monthly),
            daily_records=len(daily),
            warnings=all_warnings,
        )
        return monthly + daily, summary

    def parse_monthly(self, file_path: str, file_id: str) -> tuple[list[dict], list[str]]:
        warnings: list[str] = []
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if "Monthly details" not in wb.sheetnames:
            wb.close()
            return [], ["Sheet 'Monthly details' non trovato"]

        ws = wb["Monthly details"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            wb.close()
            return [], ["Sheet monthly troppo corto"]

        header_idx = self._detect_monthly_header_row(rows)
        date_map = self._map_monthly_columns(rows[header_idx])
        if not date_map:
            warnings.append("Nessuna colonna periodo monthly rilevata")

        # col 3 recipient / col 4 donor (0-based)
        recipient_col = 3
        donor_col = 4
        current_recipient = None
        data: list[dict] = []

        for row in rows[header_idx + 1 :]:
            if len(row) <= donor_col:
                continue

            recipient_val = row[recipient_col] if recipient_col < len(row) else None
            donor_val = row[donor_col] if donor_col < len(row) else None

            if recipient_val:
                candidate = str(recipient_val).strip()
                if self._looks_like_operator(candidate):
                    current_recipient = candidate
                    continue

            if donor_val and current_recipient:
                donor = str(donor_val).strip()
                if not self._looks_like_operator(donor):
                    continue
                if (not self.include_self_flows) and donor.upper() == current_recipient.upper():
                    continue

                for col_idx, period_date in date_map.items():
                    if col_idx >= len(row):
                        continue
                    cleaned = self._clean_value(row[col_idx])
                    if cleaned is None:
                        continue
                    data.append(
                        {
                            "file_id": file_id,
                            "period_type": "MONTHLY",
                            "period_date": period_date,
                            "donor_raw": donor,
                            "recipient_raw": current_recipient,
                            "value": cleaned,
                            "sheet_name": "Monthly details",
                            "quality_flag": "IMPUTED" if cleaned == 0 and str(row[col_idx]).strip() in {"-", ""} else "OK",
                        }
                    )

        wb.close()
        return data, warnings

    def parse_daily(self, file_path: str, file_id: str) -> tuple[list[dict], list[str]]:
        warnings: list[str] = []
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if "Daily details" not in wb.sheetnames:
            wb.close()
            return [], ["Sheet 'Daily details' non trovato"]

        ws = wb["Daily details"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 6:
            wb.close()
            return [], ["Sheet daily troppo corto"]

        header_idx = self._detect_daily_header_row(rows)
        year_hint = self._extract_year_hint(file_path)
        date_map = self._map_daily_columns(rows[header_idx], year_hint)
        if not date_map:
            warnings.append("Nessuna colonna periodo daily rilevata")

        recipient_col = 3
        donor_col = 4
        current_recipient = None
        data: list[dict] = []

        for row in rows[header_idx + 1 :]:
            if len(row) <= donor_col:
                continue

            recipient_val = row[recipient_col] if recipient_col < len(row) else None
            donor_val = row[donor_col] if donor_col < len(row) else None

            if recipient_val:
                candidate = str(recipient_val).strip()
                if self._looks_like_operator(candidate):
                    current_recipient = candidate
                    continue

            if donor_val and current_recipient:
                donor = str(donor_val).strip()
                if not self._looks_like_operator(donor):
                    continue
                if (not self.include_self_flows) and donor.upper() == current_recipient.upper():
                    continue

                for col_idx, period_date in date_map.items():
                    if col_idx >= len(row):
                        continue
                    cleaned = self._clean_value(row[col_idx])
                    if cleaned is None:
                        continue
                    data.append(
                        {
                            "file_id": file_id,
                            "period_type": "DAILY",
                            "period_date": period_date,
                            "donor_raw": donor,
                            "recipient_raw": current_recipient,
                            "value": cleaned,
                            "sheet_name": "Daily details",
                            "quality_flag": "IMPUTED" if cleaned == 0 and str(row[col_idx]).strip() in {"-", ""} else "OK",
                        }
                    )

        wb.close()
        return data, warnings

    @staticmethod
    def _detect_monthly_header_row(rows: list[tuple]) -> int:
        # trova la riga con piu token mese nei primi 12 rows
        best_idx = 4
        best_score = -1
        for idx, row in enumerate(rows[:12]):
            score = 0
            for val in row:
                if not val:
                    continue
                text = str(val).upper().strip()
                if re.match(r"^[A-Z]{3}\s*\d{2,4}$", text) or text[:3] in MNPParser.MONTH_ALIASES:
                    score += 1
            if score > best_score:
                best_idx = idx
                best_score = score
        return best_idx

    @staticmethod
    def _detect_daily_header_row(rows: list[tuple]) -> int:
        best_idx = 4
        best_score = -1
        pattern = re.compile(r"^\d{1,2}/\d{1,2}$")
        for idx, row in enumerate(rows[:12]):
            score = 0
            for val in row:
                if val and pattern.match(str(val).strip()):
                    score += 1
            if score > best_score:
                best_idx = idx
                best_score = score
        return best_idx

    def _map_monthly_columns(self, header_row: tuple) -> dict[int, date]:
        date_map: dict[int, date] = {}
        current_year: Optional[int] = None

        year_token = re.compile(r"^(20\d{2})$")
        month_year_token = re.compile(r"^([A-Za-z]{3})\s*(\d{2,4})$")

        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            text = str(cell).strip()
            upper = text.upper()

            if year_token.match(upper):
                current_year = int(upper)
                continue

            my = month_year_token.match(upper)
            if my:
                mon_key = my.group(1)[:3]
                year_txt = my.group(2)
                month = self.MONTH_ALIASES.get(mon_key)
                if not month:
                    continue
                year = int(year_txt) if len(year_txt) == 4 else (2000 + int(year_txt))
                date_map[idx] = date(year, month, 1)
                current_year = year
                continue

            mon_key = upper[:3]
            if mon_key in self.MONTH_ALIASES and current_year:
                date_map[idx] = date(current_year, self.MONTH_ALIASES[mon_key], 1)

        return date_map

    @staticmethod
    def _extract_year_hint(file_path: str) -> int:
        name = Path(file_path).name
        match = re.search(r"20\d{2}", name)
        if match:
            return int(match.group(0))
        return datetime.now().year

    def _map_daily_columns(self, header_row: tuple, start_year: int) -> dict[int, date]:
        date_map: dict[int, date] = {}
        current_year = start_year
        last_month = 0
        pattern = re.compile(r"^(\d{1,2})/(\d{1,2})$")

        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            text = str(cell).strip()
            match = pattern.match(text)
            if not match:
                continue

            day_num = int(match.group(1))
            month_num = int(match.group(2))

            # rollover anno quando si passa da dicembre a gennaio
            if last_month == 12 and month_num == 1:
                current_year += 1

            try:
                date_map[idx] = date(current_year, month_num, day_num)
                last_month = month_num
            except ValueError:
                continue

        return date_map

    @staticmethod
    def _clean_value(raw_val: object) -> float | None:
        if raw_val is None:
            return None
        if isinstance(raw_val, (int, float)):
            return float(raw_val)

        text = str(raw_val).strip()
        if text in {"", "-"}:
            return 0.0

        text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None

    @staticmethod
    def _looks_like_operator(value: str) -> bool:
        if not value:
            return False
        text = value.upper().strip()
        if text in {"RECIPIENT (1)", "RECIPIENT (2)", "RECIPIENT"}:
            return False
        if any(token in text for token in ["TREND", "TOTAL", "MONTH", "DAILY"]):
            return False
        return bool(re.match(r"^[A-Z0-9 .+]+$", text))
