"""Generic Excel template analysis, versioning and ingestion."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
import hashlib
import json
import re

import openpyxl
import pandas as pd

from mnp_cdx.db.repository import DBRepository


DATE_HINT_PATTERN = re.compile(r"(20\d{2})(\d{2})(\d{2})")
SIMPLE_DATE_PATTERN = re.compile(r"^(\d{1,2})[\-/](\d{1,2})(?:[\-/](\d{2,4}))?$")


@dataclass
class AnalyzeResult:
    workbook_signature: str
    schema: dict[str, Any]
    matched_template: dict[str, Any] | None


@dataclass
class GenericIngestResult:
    file_id: int | None
    filename: str
    checksum: str
    template_id: int
    template_name: str
    template_version: int
    created_new_template: bool
    inserted_rows: int
    skipped_duplicate: bool
    warnings: list[str]


class GenericTemplateEngine:
    def __init__(self, repo: DBRepository) -> None:
        self.repo = repo

    @staticmethod
    def checksum(file_path: str | Path) -> str:
        sha = hashlib.sha256()
        with open(file_path, "rb") as fh:
            for chunk in iter(lambda: fh.read(4096), b""):
                sha.update(chunk)
        return sha.hexdigest()

    @staticmethod
    def _file_date_hint(file_path: str | Path) -> date | None:
        name = Path(file_path).name
        match = DATE_HINT_PATTERN.search(name)
        if not match:
            return None
        try:
            return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None

    def analyze(self, file_path: str | Path) -> AnalyzeResult:
        file_path = Path(file_path)
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        schema: dict[str, Any] = {
            "engine_version": "generic-template-1.0",
            "file_date_hint": str(self._file_date_hint(file_path)) if self._file_date_hint(file_path) else None,
            "sheets": [],
        }

        for ws in wb.worksheets:
            sheet_schema = self._analyze_sheet(ws)
            schema["sheets"].append(sheet_schema)

        wb.close()

        workbook_signature = self._compute_signature(schema)
        matched = self.repo.get_template_by_signature(workbook_signature)
        return AnalyzeResult(workbook_signature=workbook_signature, schema=schema, matched_template=matched)

    def _analyze_sheet(self, ws) -> dict[str, Any]:
        max_scan_rows = min(80, ws.max_row)
        max_scan_cols = min(400, ws.max_column)

        scan_rows: list[tuple[Any, ...]] = list(
            ws.iter_rows(min_row=1, max_row=max_scan_rows, max_col=max_scan_cols, values_only=True)
        )

        header_row = self._detect_header_row(scan_rows)
        header_values = scan_rows[header_row - 1] if header_row - 1 < len(scan_rows) else ()
        headers = self._normalize_headers(header_values, max_scan_cols)

        profiles = self._profile_columns(ws, headers, data_start_row=header_row + 1, max_scan_cols=max_scan_cols)
        date_columns = [
            p["name"]
            for p in profiles
            if p["inferred_type"] == "date" or self._header_looks_date(p["name"])
        ]
        metric_columns = [
            p["name"]
            for p in profiles
            if p["inferred_type"] == "numeric" and p["non_empty_ratio"] >= 0.10
        ]

        dimension_columns = [
            p["name"]
            for p in profiles
            if p["name"] not in date_columns and p["name"] not in metric_columns
        ]

        return {
            "sheet_name": ws.title,
            "max_row": int(ws.max_row),
            "max_col": int(ws.max_column),
            "header_row": int(header_row),
            "data_start_row": int(header_row + 1),
            "columns": profiles,
            "date_columns": date_columns,
            "metric_columns": metric_columns,
            "dimension_columns": dimension_columns,
        }

    @staticmethod
    def _detect_header_row(scan_rows: list[tuple[Any, ...]]) -> int:
        best_idx = 1
        best_score = -1

        for idx, row in enumerate(scan_rows, start=1):
            non_empty = 0
            text_like = 0
            for val in row:
                if val is None:
                    continue
                sval = str(val).strip()
                if not sval:
                    continue
                non_empty += 1
                if isinstance(val, str):
                    text_like += 1
                elif any(ch.isalpha() for ch in sval):
                    text_like += 1

            # privilegia righe con testo e densita
            score = (text_like * 3) + non_empty
            if non_empty >= 3 and score > best_score:
                best_idx = idx
                best_score = score

        return best_idx

    @staticmethod
    def _clean_header_text(raw: Any, fallback_idx: int) -> str:
        if raw is None:
            return f"col_{fallback_idx}"
        txt = str(raw).strip().replace("\n", " ")
        txt = re.sub(r"\s+", " ", txt)
        if not txt:
            return f"col_{fallback_idx}"
        return txt[:120]

    def _normalize_headers(self, header_values: tuple[Any, ...], max_scan_cols: int) -> list[str]:
        headers: list[str] = []
        used: dict[str, int] = {}

        for idx in range(1, max_scan_cols + 1):
            raw = header_values[idx - 1] if idx - 1 < len(header_values) else None
            name = self._clean_header_text(raw, idx)

            if name in used:
                used[name] += 1
                name = f"{name}__{used[name]}"
            else:
                used[name] = 1

            headers.append(name)

        return headers

    def _profile_columns(
        self,
        ws,
        headers: list[str],
        data_start_row: int,
        max_scan_cols: int,
        max_sample_rows: int = 600,
    ) -> list[dict[str, Any]]:
        stats = [
            {
                "non_empty": 0,
                "numeric": 0,
                "date": 0,
                "text": 0,
                "samples": [],
            }
            for _ in range(max_scan_cols)
        ]

        sampled = 0
        for row in ws.iter_rows(
            min_row=data_start_row,
            max_row=min(ws.max_row, data_start_row + max_sample_rows - 1),
            max_col=max_scan_cols,
            values_only=True,
        ):
            sampled += 1
            for col_idx in range(max_scan_cols):
                val = row[col_idx] if col_idx < len(row) else None
                if val is None:
                    continue

                sval = str(val).strip()
                if not sval:
                    continue

                st = stats[col_idx]
                st["non_empty"] += 1

                t = self._infer_value_type(val)
                st[t] += 1

                if len(st["samples"]) < 3:
                    st["samples"].append(sval[:80])

        sampled = max(sampled, 1)
        profiles: list[dict[str, Any]] = []
        for col_idx in range(max_scan_cols):
            st = stats[col_idx]
            inferred = "text"
            if st["date"] >= st["numeric"] and st["date"] >= st["text"] and st["date"] > 0:
                inferred = "date"
            elif st["numeric"] >= st["text"] and st["numeric"] > 0:
                inferred = "numeric"

            non_empty_ratio = st["non_empty"] / sampled
            if non_empty_ratio < 0.01:
                # colonne praticamente vuote: le ignoriamo in template operativo
                continue

            profiles.append(
                {
                    "index": col_idx + 1,
                    "name": headers[col_idx],
                    "inferred_type": inferred,
                    "non_empty_ratio": round(non_empty_ratio, 4),
                    "samples": st["samples"],
                }
            )

        return profiles

    @staticmethod
    def _infer_value_type(value: Any) -> str:
        if isinstance(value, (datetime, date)):
            return "date"
        if isinstance(value, (int, float)):
            return "numeric"

        sval = str(value).strip()
        if not sval:
            return "text"

        # numero stringa
        num_candidate = sval.replace(" ", "").replace(",", ".")
        try:
            float(num_candidate)
            return "numeric"
        except ValueError:
            pass

        # data stringa d/m[/y]
        if SIMPLE_DATE_PATTERN.match(sval):
            return "date"

        try:
            datetime.fromisoformat(sval)
            return "date"
        except ValueError:
            return "text"

    @staticmethod
    def _header_looks_date(header_name: str) -> bool:
        h = header_name.lower()
        tokens = ["date", "data", "day", "month", "year", "period"]
        return any(token in h for token in tokens)

    @staticmethod
    def _compute_signature(schema: dict[str, Any]) -> str:
        payload = {
            "engine_version": schema.get("engine_version"),
            "sheets": [],
        }

        for sheet in schema.get("sheets", []):
            payload["sheets"].append(
                {
                    "sheet_name": sheet.get("sheet_name", "").lower(),
                    "header_row": sheet.get("header_row"),
                    "columns": [col.get("name") for col in sheet.get("columns", [])],
                    "metric_columns": sheet.get("metric_columns", []),
                    "date_columns": sheet.get("date_columns", []),
                }
            )

        raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
        return hashlib.sha256(raw.encode("utf-8")).hexdigest()

    def ingest(
        self,
        file_path: str | Path,
        template_id: int | None = None,
        template_name: str | None = None,
        force: bool = False,
    ) -> GenericIngestResult:
        file_path = Path(file_path)
        checksum = self.checksum(file_path)

        if self.repo.file_exists(checksum):
            if not force:
                # se file duplicato, serve almeno template info per response
                analysis = self.analyze(file_path)
                matched = analysis.matched_template
                if matched is None:
                    matched, _ = self.repo.create_or_reuse_template(
                        signature=analysis.workbook_signature,
                        schema=analysis.schema,
                        template_name=template_name,
                    )

                return GenericIngestResult(
                    file_id=None,
                    filename=file_path.name,
                    checksum=checksum,
                    template_id=int(matched["template_id"]),
                    template_name=str(matched["template_name"]),
                    template_version=int(matched["template_version"]),
                    created_new_template=False,
                    inserted_rows=0,
                    skipped_duplicate=True,
                    warnings=["File gia ingestito (checksum duplicate)"] ,
                )
            self.repo.delete_file_everywhere_by_checksum(checksum)

        analysis = self.analyze(file_path)

        if template_id is not None:
            template = self.repo.get_template_by_id(template_id)
            if template is None:
                raise ValueError(f"Template id non trovato: {template_id}")
            created_new_template = False
        else:
            template, created_new_template = self.repo.create_or_reuse_template(
                signature=analysis.workbook_signature,
                schema=analysis.schema,
                template_name=template_name,
            )

        file_id = self.repo.insert_ingest_file(
            filename=file_path.name,
            checksum=checksum,
            parser_version="generic-template-1.0",
        )

        template_schema = template["schema"]
        rows_inserted, warnings = self._materialize_rows(file_path, file_id, int(template["template_id"]), template_schema)

        self.repo.update_ingest_status(file_id, "OK", rows_inserted)
        self.repo.insert_excel_ingest_event(
            file_id=file_id,
            template_id=int(template["template_id"]),
            status="OK",
            row_count=rows_inserted,
            notes="; ".join(warnings) if warnings else None,
        )

        return GenericIngestResult(
            file_id=file_id,
            filename=file_path.name,
            checksum=checksum,
            template_id=int(template["template_id"]),
            template_name=str(template["template_name"]),
            template_version=int(template["template_version"]),
            created_new_template=created_new_template,
            inserted_rows=rows_inserted,
            skipped_duplicate=False,
            warnings=warnings,
        )

    def _materialize_rows(
        self,
        file_path: Path,
        file_id: int,
        template_id: int,
        template_schema: dict[str, Any],
    ) -> tuple[int, list[str]]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        warnings: list[str] = []
        file_date_hint = self._file_date_hint(file_path)

        chunk: list[dict[str, Any]] = []
        inserted_total = 0

        for sheet_cfg in template_schema.get("sheets", []):
            sheet_name = sheet_cfg.get("sheet_name")
            if sheet_name not in wb.sheetnames:
                warnings.append(f"Sheet assente in file: {sheet_name}")
                continue

            ws = wb[sheet_name]
            columns = sheet_cfg.get("columns", [])
            if not columns:
                continue

            date_cols = set(sheet_cfg.get("date_columns", []))
            metric_cols = set(sheet_cfg.get("metric_columns", []))
            dimension_cols = set(sheet_cfg.get("dimension_columns", []))

            max_col = max(int(col["index"]) for col in columns)
            data_start_row = int(sheet_cfg.get("data_start_row", 2))

            for row_number, row in enumerate(
                ws.iter_rows(min_row=data_start_row, max_col=max_col, values_only=True),
                start=data_start_row,
            ):
                row_map: dict[str, Any] = {}
                has_data = False

                for col in columns:
                    idx = int(col["index"]) - 1
                    col_name = str(col["name"])
                    val = row[idx] if idx < len(row) else None
                    norm = self._normalize_scalar(val)
                    row_map[col_name] = norm
                    if norm not in (None, ""):
                        has_data = True

                if not has_data:
                    continue

                event_date = self._extract_event_date(row_map, date_cols, file_date_hint)

                metrics: dict[str, float] = {}
                dimensions: dict[str, Any] = {}

                for key in metric_cols:
                    num = self._to_float(row_map.get(key))
                    if num is not None:
                        metrics[key] = num

                for key in dimension_cols:
                    val = row_map.get(key)
                    if val not in (None, ""):
                        dimensions[key] = val

                if not metrics and not dimensions:
                    continue

                chunk.append(
                    {
                        "file_id": file_id,
                        "template_id": template_id,
                        "sheet_name": sheet_name,
                        "row_number": row_number,
                        "event_date": event_date,
                        "metrics_json": json.dumps(metrics, ensure_ascii=False),
                        "dimensions_json": json.dumps(dimensions, ensure_ascii=False),
                        "raw_json": json.dumps(row_map, ensure_ascii=False),
                    }
                )

                if len(chunk) >= 5000:
                    inserted_total += self.repo.insert_generic_row_dataframe(pd.DataFrame(chunk))
                    chunk = []

        if chunk:
            inserted_total += self.repo.insert_generic_row_dataframe(pd.DataFrame(chunk))

        wb.close()
        return inserted_total, warnings

    @staticmethod
    def _normalize_scalar(value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float):
            return float(value)
        if isinstance(value, int):
            return int(value)
        text = str(value).strip()
        return text if text else None

    def _extract_event_date(self, row_map: dict[str, Any], date_cols: set[str], fallback: date | None) -> date | None:
        for col_name in date_cols:
            raw = row_map.get(col_name)
            if raw is None:
                continue
            parsed = self._parse_date(raw)
            if parsed:
                return parsed
        return fallback

    @staticmethod
    def _parse_date(raw: Any) -> date | None:
        if raw is None:
            return None
        if isinstance(raw, date) and not isinstance(raw, datetime):
            return raw
        if isinstance(raw, datetime):
            return raw.date()

        text = str(raw).strip()
        if not text:
            return None

        try:
            return datetime.fromisoformat(text).date()
        except ValueError:
            pass

        match = SIMPLE_DATE_PATTERN.match(text)
        if match:
            day = int(match.group(1))
            month = int(match.group(2))
            year_raw = match.group(3)
            if not year_raw:
                return None
            year = int(year_raw)
            if year < 100:
                year += 2000
            try:
                return date(year, month, day)
            except ValueError:
                return None

        return None

    @staticmethod
    def _to_float(raw: Any) -> float | None:
        if raw is None:
            return None
        if isinstance(raw, (int, float)):
            return float(raw)

        txt = str(raw).strip().replace(" ", "")
        if not txt:
            return None
        txt = txt.replace(",", ".")
        try:
            return float(txt)
        except ValueError:
            return None
